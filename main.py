"""
FastAPI Job Scraper Application

This application provides endpoints to scrape job postings from Indeed and LinkedIn
using browser automation with Playwright.

Setup Instructions:
1. Install dependencies: pip install -r requirements.txt
2. Install Playwright browsers: playwright install chromium
3. Set up environment variables (optional):
   - LINKEDIN_EMAIL: Your LinkedIn email
   - LINKEDIN_PASSWORD: Your LinkedIn password
4. Run the application: uvicorn main:app --reload

Usage:
- GET /scrape/indeed?job=python+developer&location=remote
- GET /scrape/linkedin?job=python+developer&location=remote
"""

import asyncio
import logging
from typing import List, Optional
from fastapi import FastAPI, HTTPException, Query, File, UploadFile
from fastapi.responses import FileResponse
from fastapi.middleware.cors import CORSMiddleware
from pydantic import BaseModel
from dotenv import load_dotenv
import os
from datetime import datetime

from scrapers.indeed_scraper import IndeedScraper
from scrapers.linkedin_scraper import LinkedInScraper
from scrapers.enhanced_linkedin_scraper import EnhancedLinkedInScraper
from scrapers.google_jobs_scraper import GoogleJobsScraper
from utils.browser_manager import BrowserManager
from utils.excel_exporter import ExcelExporter
from models import JobPosting, ScrapingResponse

# Load environment variables
load_dotenv()

# Configure logging
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

# Reduce noisy logs from scrapers; keep API logs at INFO
for _name in [
    'scrapers.linkedin_scraper_playwright',
    'scrapers.company_recon_scraper',
    'scrapers.enhanced_linkedin_scraper',
    'scrapers.linkedin_scraper',
]:
    logging.getLogger(_name).setLevel(logging.WARNING)

app = FastAPI(
    title="Job Scraper API",
    description="API for scraping job postings from Indeed and LinkedIn",
    version="1.0.0"
)

# Add CORS middleware to allow any origin
app.add_middleware(
    CORSMiddleware,
    allow_origins=[
        "http://localhost:3000",
        "http://localhost:8000", 
        "https://admin.workbrook.us",
        "https://lead.workbrook.us",
        "https://workbrook.us",
        "https://*.workbrook.us"
    ],
    allow_credentials=True,
    allow_methods=["GET", "POST", "PUT", "DELETE", "OPTIONS"],
    allow_headers=["*"],
    expose_headers=["Content-Disposition"],
    max_age=3600,
)

@app.get("/", response_model=dict)
async def root():
    """Root endpoint with API information"""
    return {
        "message": "Job Scraper API",
        "version": "1.0.0",
        "endpoints": {
            "indeed_json": "/scrape/indeed?job=<job_title>&location=<location>&max_jobs=5",
            "indeed_excel": "/scrape/indeed/excel?job=<job_title>&location=<location>&max_jobs=5",
            "linkedin_excel": "/scrape/linkedin?job=<job_title>&location=<location>&max_jobs=5",
            "google_jobs_json": "/scrape/google-jobs?job=<job_title>&location=<location>&max_jobs=5",
            "google_jobs_excel": "/scrape/google-jobs/excel?job=<job_title>&location=<location>&max_jobs=5",
            "list_exports": "/exports",
            "download_export": "/exports/{filename}"
        },
        "default_max_jobs": 5,
        "max_jobs_limit": 20
    }

@app.get("/scrape/indeed", response_model=ScrapingResponse)
async def scrape_indeed(
    job: str = Query(..., description="Job title to search for"),
    location: str = Query(..., description="Job location"),
    max_jobs: int = Query(5, description="Maximum number of jobs to scrape", ge=1, le=20)
):
    """
    Scrape job postings from Indeed
    
    Args:
        job: Job title to search for
        location: Job location
        max_jobs: Maximum number of jobs to scrape (default: 5, max: 20)
    
    Returns:
        ScrapingResponse with job listings
    """
    try:
        logger.info(f"Starting Indeed scrape for: {job} in {location}")
        
        # Use browser manager as async context manager
        async with BrowserManager() as browser_manager:
            # Create scraper instance
            scraper = IndeedScraper(browser_manager)
            
            # Scrape jobs
            jobs = await scraper.scrape_jobs(job, location, max_jobs)
            
            return ScrapingResponse(
                success=True,
                message=f"Successfully scraped {len(jobs)} jobs from Indeed",
                jobs=jobs,
                total_count=len(jobs)
            )
        
    except Exception as e:
        logger.error(f"Error scraping Indeed: {str(e)}")
        raise HTTPException(
            status_code=500,
            detail=f"Failed to scrape Indeed: {str(e)}"
        )

@app.get("/scrape/linkedin")
async def scrape_linkedin(
    job: str = Query(..., description="Job title to search for"),
    location: str = Query(..., description="Job location"),
    max_jobs: int = Query(100, description="Maximum number of jobs to scrape", ge=1, le=200)
):
    """
    Enhanced LinkedIn scraper with company contact reconnaissance fallback
    
    This endpoint scrapes job postings from LinkedIn and enhances them with company
    contact information using reconnaissance techniques when LinkedIn doesn't provide
    email addresses.
    
    Args:
        job: Job title to search for
        location: Job location
        max_jobs: Maximum number of jobs to scrape (default: 5, max: 20)
    
    Returns:
        Excel file download with enhanced contact information
    """
    try:
        logger.info(f"Starting enhanced LinkedIn scrape for: {job} in {location}")
        
        # Create enhanced scraper instance with reconnaissance fallback
        scraper = EnhancedLinkedInScraper()
        
        # Scrape jobs with company contact enhancement
        jobs = await scraper.scrape_jobs_with_contacts(job, location, max_jobs)
        
        if not jobs:
            raise HTTPException(
                status_code=404,
                detail="No jobs found for the given criteria"
            )
        
        # Create Excel file in memory and return as download
        import io
        import pandas as pd
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
        from datetime import datetime
        
        # Create DataFrame with job data
        data = []
        for i, job_obj in enumerate(jobs, 1):
            data.append({
                'ID': i,
                'Job Title': job_obj.title,
                'Company': job_obj.company,
                'Location': job_obj.location,
                'Description': job_obj.description or '',
                'Poster Name': job_obj.poster_name or '',
                'Poster Position': job_obj.poster_position or '',
                'Email': job_obj.email or '',
                'Apply Link': job_obj.url or '',
                'Date Posted': job_obj.date_posted or '',
                'Scraped Date': datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            })
        
        df = pd.DataFrame(data)
        
        # Create Excel file in memory
        output = io.BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name='Jobs', index=False)
            
            # Get the workbook and worksheet
            workbook = writer.book
            worksheet = writer.sheets['Jobs']
            
            # Apply formatting
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            header_alignment = Alignment(horizontal="center", vertical="center")
            
            # Apply header formatting
            for col in range(1, len(df.columns) + 1):
                cell = worksheet.cell(row=1, column=col)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
                
            # Auto-adjust column widths
            for col in range(1, len(df.columns) + 1):
                column_letter = get_column_letter(col)
                max_length = 0
                
                # Find the maximum length in the column
                for row in range(1, worksheet.max_row + 1):
                    cell_value = worksheet[f"{column_letter}{row}"].value
                    if cell_value:
                        max_length = max(max_length, len(str(cell_value)))
                        
                # Set column width (with some padding)
                adjusted_width = min(max_length + 2, 50)
                worksheet.column_dimensions[column_letter].width = adjusted_width
                
            # Add borders to all cells
            thin_border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            for row in worksheet.iter_rows(min_row=1, max_row=worksheet.max_row, min_col=1, max_col=len(df.columns)):
                for cell in row:
                    cell.border = thin_border
        
        # Prepare the response
        output.seek(0)
        
        # Generate filename
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        safe_job_title = job.replace(' ', '_').replace('/', '_').replace('\\', '_')[:30]
        safe_location = location.replace(' ', '_').replace('/', '_').replace('\\', '_')[:30]
        filename = f"linkedin_{safe_job_title}_{safe_location}_{timestamp}.xlsx"
        
        # Return the Excel file as download
        from fastapi.responses import StreamingResponse
        return StreamingResponse(
            io.BytesIO(output.getvalue()),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename={filename}"}
        )
    
    except Exception as e:
        logger.error(f"Error scraping enhanced LinkedIn: {str(e)}")
        raise HTTPException(
            status_code=500,
            detail=f"Failed to scrape enhanced LinkedIn: {str(e)}"
        )

@app.get("/scrape/linkedin-enhanced")
async def scrape_linkedin_enhanced(
    job: str = Query(..., description="Job title to search for"),
    location: str = Query(..., description="Job location"),
    max_jobs: int = Query(50, description="Maximum number of jobs to scrape", ge=1, le=200)
):
    """
    Enhanced LinkedIn scraper that also extracts company contact information
    
    Args:
        job: Job title to search for
        location: Job location
        max_jobs: Maximum number of jobs to scrape (default: 5, max: 20)
    
    Returns:
        Excel file download with enhanced contact information
    """
    try:
        logger.info(f"Starting enhanced LinkedIn scrape for: {job} in {location}")
        
        # Create enhanced scraper instance
        scraper = EnhancedLinkedInScraper()
        
        # Scrape jobs with company contact enhancement
        jobs = await scraper.scrape_jobs_with_contacts(job, location, max_jobs)
        
        if not jobs:
            raise HTTPException(
                status_code=404,
                detail="No jobs found for the given criteria"
            )
        
        # Create Excel file in memory and return as download
        import io
        import pandas as pd
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
        from datetime import datetime
        
        # Create DataFrame with enhanced job data
        data = []
        for i, job_obj in enumerate(jobs, 1):
            data.append({
                'ID': i,
                'Job Title': job_obj.title,
                'Company': job_obj.company,
                'Location': job_obj.location,
                'Description': job_obj.description or '',
                'Poster Name': job_obj.poster_name or '',
                'Poster Position': job_obj.poster_position or '',
                'Email': job_obj.email or '',
                'Apply Link': job_obj.url or '',
                'Date Posted': job_obj.date_posted or '',
                'Scraped Date': datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            })
        
        df = pd.DataFrame(data)
        
        # Create Excel file in memory
        output = io.BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name='Jobs', index=False)
            
            # Get the workbook and worksheet
            workbook = writer.book
            worksheet = writer.sheets['Jobs']
            
            # Apply formatting
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            header_alignment = Alignment(horizontal="center", vertical="center")
            
            # Apply header formatting
            for col in range(1, len(df.columns) + 1):
                cell = worksheet.cell(row=1, column=col)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
                
            # Auto-adjust column widths
            for col in range(1, len(df.columns) + 1):
                column_letter = get_column_letter(col)
                max_length = 0
                
                # Find the maximum length in the column
                for row in range(1, worksheet.max_row + 1):
                    cell_value = worksheet[f"{column_letter}{row}"].value
                    if cell_value:
                        max_length = max(max_length, len(str(cell_value)))
                        
                # Set column width (with some padding)
                adjusted_width = min(max_length + 2, 50)
                worksheet.column_dimensions[column_letter].width = adjusted_width
                
            # Add borders to all cells
            thin_border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            for row in worksheet.iter_rows(min_row=1, max_row=worksheet.max_row, min_col=1, max_col=len(df.columns)):
                for cell in row:
                    cell.border = thin_border
        
        # Prepare the response
        output.seek(0)
        
        # Generate filename
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        safe_job_title = job.replace(' ', '_').replace('/', '_').replace('\\', '_')[:30]
        safe_location = location.replace(' ', '_').replace('/', '_').replace('\\', '_')[:30]
        filename = f"linkedin_enhanced_{safe_job_title}_{safe_location}_{timestamp}.xlsx"
        
        # Return the Excel file as download
        from fastapi.responses import StreamingResponse
        return StreamingResponse(
            io.BytesIO(output.getvalue()),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename={filename}"}
        )
    
    except Exception as e:
        logger.error(f"Error scraping enhanced LinkedIn: {str(e)}")
        raise HTTPException(
            status_code=500,
            detail=f"Failed to scrape enhanced LinkedIn: {str(e)}"
        )

@app.get("/health")
async def health_check():
    """Health check endpoint"""
    return {"status": "healthy", "message": "Job Scraper API is running"}

@app.get("/scrape/indeed/excel")
async def scrape_indeed_excel(
    job: str = Query(..., description="Job title to search for"),
    location: str = Query(..., description="Job location"),
    max_jobs: int = Query(5, description="Maximum number of jobs to scrape", ge=1, le=20)
):
    """
    Scrape job postings from Indeed and download as Excel file
    
    Args:
        job: Job title to search for
        location: Job location
        max_jobs: Maximum number of jobs to scrape (default: 5, max: 20)
    
    Returns:
        Excel file download
    """
    try:
        logger.info(f"Starting Indeed scrape for Excel: {job} in {location}")
        
        # Use browser manager as async context manager
        async with BrowserManager() as browser_manager:
            # Create scraper instance
            scraper = IndeedScraper(browser_manager)
            
            # Scrape jobs
            jobs = await scraper.scrape_jobs(job, location, max_jobs)
            
            if not jobs:
                raise HTTPException(
                    status_code=404,
                    detail="No jobs found for the given criteria"
                )
            
            # Create Excel file in memory and return as download
            import io
            import pandas as pd
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            from openpyxl.utils import get_column_letter
            
            # Create DataFrame with job data
            data = []
            for i, job_obj in enumerate(jobs, 1):
                data.append({
                    'ID': i,
                    'Job Title': job_obj.title,
                    'Company': job_obj.company,
                    'Location': job_obj.location,
                    'Description': job_obj.description or '',
                    'Poster Name': job_obj.poster_name or '',
                    'Poster Position': job_obj.poster_position or '',
                    'Email': job_obj.email or '',
                    'Apply Link': job_obj.url or '',
                    'Date Posted': job_obj.date_posted or '',
                    'Scraped Date': datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                })
            
            df = pd.DataFrame(data)
            
            # Create Excel file in memory
            output = io.BytesIO()
            with pd.ExcelWriter(output, engine='openpyxl') as writer:
                df.to_excel(writer, sheet_name='Jobs', index=False)
                
                # Get the workbook and worksheet
                workbook = writer.book
                worksheet = writer.sheets['Jobs']
                
                # Apply formatting
                header_font = Font(bold=True, color="FFFFFF")
                header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                header_alignment = Alignment(horizontal="center", vertical="center")
                
                # Apply header formatting
                for col in range(1, len(df.columns) + 1):
                    cell = worksheet.cell(row=1, column=col)
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = header_alignment
                    
                # Auto-adjust column widths
                for col in range(1, len(df.columns) + 1):
                    column_letter = get_column_letter(col)
                    max_length = 0
                    
                    # Find the maximum length in the column
                    for row in range(1, worksheet.max_row + 1):
                        cell_value = worksheet[f"{column_letter}{row}"].value
                        if cell_value:
                            max_length = max(max_length, len(str(cell_value)))
                            
                    # Set column width (with some padding)
                    adjusted_width = min(max_length + 2, 50)
                    worksheet.column_dimensions[column_letter].width = adjusted_width
                    
                # Add borders to all cells
                thin_border = Border(
                    left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='thin'),
                    bottom=Side(style='thin')
                )
                
                for row in worksheet.iter_rows(min_row=1, max_row=worksheet.max_row, min_col=1, max_col=len(df.columns)):
                    for cell in row:
                        cell.border = thin_border
            
            # Prepare the response
            output.seek(0)
            
            # Generate filename
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            safe_job_title = job.replace(' ', '_').replace('/', '_').replace('\\', '_')[:30]
            safe_location = location.replace(' ', '_').replace('/', '_').replace('\\', '_')[:30]
            filename = f"indeed_{safe_job_title}_{safe_location}_{timestamp}.xlsx"
            
            # Return the Excel file as download
            from fastapi.responses import StreamingResponse
            return StreamingResponse(
                io.BytesIO(output.getvalue()),
                media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                headers={"Content-Disposition": f"attachment; filename={filename}"}
            )
        
    except Exception as e:
        logger.error(f"Error scraping Indeed for Excel: {str(e)}")
        raise HTTPException(
            status_code=500,
            detail=f"Failed to scrape Indeed: {str(e)}"
        )

@app.get("/scrape/google-jobs", response_model=ScrapingResponse)
async def scrape_google_jobs(
    job: str = Query(..., description="Job title to search for"),
    location: str = Query(..., description="Job location"),
    max_jobs: int = Query(5, description="Maximum number of jobs to scrape", ge=1, le=20)
):
    """
    Scrape job postings from Google Jobs
    
    Args:
        job: Job title to search for
        location: Job location
        max_jobs: Maximum number of jobs to scrape (default: 5, max: 20)
    
    Returns:
        ScrapingResponse with job listings
    """
    try:
        logger.info(f"Starting Google Jobs scrape for: {job} in {location}")
        
        # Use browser manager as async context manager
        async with BrowserManager() as browser_manager:
            # Create scraper instance
            scraper = GoogleJobsScraper(browser_manager)
            
            # Scrape jobs
            jobs = await scraper.scrape_jobs(job, location, max_jobs)
            
            return ScrapingResponse(
                success=True,
                message=f"Successfully scraped {len(jobs)} jobs from Google Jobs",
                jobs=jobs,
                total_count=len(jobs)
            )
        
    except Exception as e:
        logger.error(f"Error scraping Google Jobs: {str(e)}")
        raise HTTPException(
            status_code=500,
            detail=f"Failed to scrape Google Jobs: {str(e)}"
        )

@app.get("/scrape/google")
async def scrape_google(
    job: str = Query(..., description="Job title to search for"),
    location: str = Query(..., description="Job location"),
    max_jobs: int = Query(5, description="Maximum number of jobs to scrape", ge=1, le=20)
):
    """
    Scrape job postings from Google Jobs and download as Excel file
    
    Args:
        job: Job title to search for
        location: Job location
        max_jobs: Maximum number of jobs to scrape (default: 5, max: 20)
    
    Returns:
        Excel file download
    """
    try:
        logger.info(f"Starting Google Jobs scrape for Excel: {job} in {location}")
        
        # Use browser manager as async context manager
        async with BrowserManager() as browser_manager:
            # Create scraper instance
            scraper = GoogleJobsScraper(browser_manager)
            
            # Scrape jobs
            jobs = await scraper.scrape_jobs(job, location, max_jobs)
            
            if not jobs:
                raise HTTPException(
                    status_code=404,
                    detail="No jobs found for the given criteria"
                )
            
            # Create Excel file in memory and return as download
            import io
            import pandas as pd
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            from openpyxl.utils import get_column_letter
            
            # Create DataFrame with job data
            data = []
            for i, job_obj in enumerate(jobs, 1):
                data.append({
                    'ID': i,
                    'Job Title': job_obj.title,
                    'Company': job_obj.company,
                    'Location': job_obj.location,
                    'Description': job_obj.description or '',
                    'Poster Name': job_obj.poster_name or '',
                    'Poster Position': job_obj.poster_position or '',
                    'Email': job_obj.email or '',
                    'Apply Link': job_obj.url or '',
                    'Date Posted': job_obj.date_posted or '',
                    'Scraped Date': datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                })
            
            df = pd.DataFrame(data)
            
            # Create Excel file in memory
            output = io.BytesIO()
            with pd.ExcelWriter(output, engine='openpyxl') as writer:
                df.to_excel(writer, sheet_name='Jobs', index=False)
                
                # Get the workbook and worksheet
                workbook = writer.book
                worksheet = writer.sheets['Jobs']
                
                # Apply formatting
                header_font = Font(bold=True, color="FFFFFF")
                header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                header_alignment = Alignment(horizontal="center", vertical="center")
                
                # Apply header formatting
                for col in range(1, len(df.columns) + 1):
                    cell = worksheet.cell(row=1, column=col)
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = header_alignment
                    
                # Auto-adjust column widths
                for col in range(1, len(df.columns) + 1):
                    column_letter = get_column_letter(col)
                    max_length = 0
                    
                    # Find the maximum length in the column
                    for row in range(1, worksheet.max_row + 1):
                        cell_value = worksheet[f"{column_letter}{row}"].value
                        if cell_value:
                            max_length = max(max_length, len(str(cell_value)))
                            
                    # Set column width (with some padding)
                    adjusted_width = min(max_length + 2, 50)
                    worksheet.column_dimensions[column_letter].width = adjusted_width
                    
                # Add borders to all cells
                thin_border = Border(
                    left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='thin'),
                    bottom=Side(style='thin')
                )
                
                for row in worksheet.iter_rows(min_row=1, max_row=worksheet.max_row, min_col=1, max_col=len(df.columns)):
                    for cell in row:
                        cell.border = thin_border
            
            # Prepare the response
            output.seek(0)
            
            # Generate filename
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            safe_job_title = job.replace(' ', '_').replace('/', '_').replace('\\', '_')[:30]
            safe_location = location.replace(' ', '_').replace('/', '_').replace('\\', '_')[:30]
            filename = f"google_jobs_{safe_job_title}_{safe_location}_{timestamp}.xlsx"
            
            # Return the Excel file as download
            from fastapi.responses import StreamingResponse
            return StreamingResponse(
                io.BytesIO(output.getvalue()),
                media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                headers={"Content-Disposition": f"attachment; filename={filename}"}
            )
        
    except Exception as e:
        logger.error(f"Error scraping Google Jobs for Excel: {str(e)}")
        raise HTTPException(
            status_code=500,
            detail=f"Failed to scrape Google Jobs: {str(e)}"
        )

@app.get("/exports")
async def list_exports():
    """List all exported Excel files"""
    try:
        exporter = ExcelExporter()
        exports = exporter.list_exports()
        return {
            "exports": exports,
            "total_files": len(exports),
            "export_directory": exporter.get_export_directory()
        }
    except Exception as e:
        logger.error(f"Error listing exports: {str(e)}")
        raise HTTPException(
            status_code=500,
            detail=f"Failed to list exports: {str(e)}"
        )

@app.get("/exports/{filename}")
async def download_export(filename: str):
    """Download a specific exported Excel file"""
    try:
        exporter = ExcelExporter()
        file_path = exporter.output_dir / filename
        
        if not file_path.exists():
            raise HTTPException(
                status_code=404,
                detail=f"File {filename} not found"
            )
        
        return FileResponse(
            path=str(file_path),
            filename=filename,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error downloading export {filename}: {str(e)}")
        raise HTTPException(
            status_code=500,
            detail=f"Failed to download file: {str(e)}"
        )

if __name__ == "__main__":
    import uvicorn
    uvicorn.run(app, host="0.0.0.0", port=8000) 