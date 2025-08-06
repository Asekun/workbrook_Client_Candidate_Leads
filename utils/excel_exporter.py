"""
Excel Exporter for Job Scraping Data

This module provides functionality to export scraped job data to Excel files
with proper formatting and organization.
"""

import pandas as pd
import os
from datetime import datetime
from typing import List, Dict, Any
from pathlib import Path
import logging

from models import JobPosting

logger = logging.getLogger(__name__)

class ExcelExporter:
    """
    Exports job scraping data to Excel files with proper formatting.
    
    This class handles the creation of Excel files with job data,
    including proper column formatting, headers, and file organization.
    """
    
    def __init__(self, output_dir: str = "exports"):
        """
        Initialize the Excel exporter.
        
        Args:
            output_dir: Directory to save Excel files
        """
        self.output_dir = Path(output_dir)
        self.output_dir.mkdir(exist_ok=True)
        
    def export_jobs_to_excel(self, jobs: List[JobPosting], source: str, job_title: str, location: str) -> str:
        """
        Export job listings to an Excel file.
        
        Args:
            jobs: List of JobPosting objects
            source: Source website (e.g., 'indeed', 'linkedin')
            job_title: Job title that was searched
            location: Location that was searched
            
        Returns:
            Path to the created Excel file
        """
        try:
            # Convert jobs to DataFrame
            df = self._prepare_dataframe(jobs)
            
            # Generate filename
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            safe_job_title = self._sanitize_filename(job_title)
            safe_location = self._sanitize_filename(location)
            filename = f"{source}_{safe_job_title}_{safe_location}_{timestamp}.xlsx"
            filepath = self.output_dir / filename
            
            # Create Excel writer with formatting
            with pd.ExcelWriter(filepath, engine='openpyxl') as writer:
                # Write main data
                df.to_excel(writer, sheet_name='Jobs', index=False)
                
                # Get the workbook and worksheet
                workbook = writer.book
                worksheet = writer.sheets['Jobs']
                
                # Apply formatting
                self._apply_excel_formatting(worksheet, df)
                
                # Add summary sheet
                self._add_summary_sheet(workbook, jobs, source, job_title, location)
            
            logger.info(f"Successfully exported {len(jobs)} jobs to {filepath}")
            return str(filepath)
            
        except Exception as e:
            logger.error(f"Error exporting jobs to Excel: {str(e)}")
            raise
            
    def _prepare_dataframe(self, jobs: List[JobPosting]) -> pd.DataFrame:
        """
        Prepare job data for DataFrame conversion.
        
        Args:
            jobs: List of JobPosting objects
            
        Returns:
            Pandas DataFrame with job data
        """
        data = []
        for i, job in enumerate(jobs, 1):
            data.append({
                'ID': i,
                'Job Title': job.title,
                'Company': job.company,
                'Location': job.location,
                'URL': job.url or '',
                'Scraped Date': datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            })
            
        return pd.DataFrame(data)
        
    def _apply_excel_formatting(self, worksheet, df: pd.DataFrame):
        """
        Apply formatting to the Excel worksheet.
        
        Args:
            worksheet: OpenPyXL worksheet object
            df: Pandas DataFrame
        """
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
        
        # Header formatting
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        # Apply header formatting
        for col in range(1, len(df.columns) + 1):
            cell = worksheet.cell(row=1, column=col)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            
        # Auto-adjust column widths
        for col in range(1, len(df.columns) + 1):
            column_letter = get_column_letter(col)
            max_length = 0
            
            # Find the maximum length in the column
            for row in range(1, worksheet.max_row + 1):
                cell_value = worksheet[f"{column_letter}{row}"].value
                if cell_value:
                    max_length = max(max_length, len(str(cell_value)))
                    
            # Set column width (with some padding)
            adjusted_width = min(max_length + 2, 50)
            worksheet.column_dimensions[column_letter].width = adjusted_width
            
        # Add borders to all cells
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        for row in worksheet.iter_rows(min_row=1, max_row=worksheet.max_row, min_col=1, max_col=len(df.columns)):
            for cell in row:
                cell.border = thin_border
                
    def _add_summary_sheet(self, workbook, jobs: List[JobPosting], source: str, job_title: str, location: str):
        """
        Add a summary sheet to the Excel workbook.
        
        Args:
            workbook: OpenPyXL workbook object
            jobs: List of JobPosting objects
            source: Source website
            job_title: Job title that was searched
            location: Location that was searched
        """
        from openpyxl.styles import Font, PatternFill, Alignment
        
        # Create summary worksheet
        summary_ws = workbook.create_sheet("Summary", 0)
        
        # Summary data
        summary_data = [
            ["Scraping Summary", ""],
            ["", ""],
            ["Source Website", source.upper()],
            ["Job Title Searched", job_title],
            ["Location Searched", location],
            ["Total Jobs Found", len(jobs)],
            ["Scraping Date", datetime.now().strftime("%Y-%m-%d %H:%M:%S")],
            ["", ""],
            ["Company Distribution", ""],
        ]
        
        # Add company distribution
        if jobs:
            company_counts = {}
            for job in jobs:
                company_counts[job.company] = company_counts.get(job.company, 0) + 1
                
            for company, count in sorted(company_counts.items(), key=lambda x: x[1], reverse=True):
                summary_data.append([company, count])
                
        # Write summary data
        for row_idx, (col1, col2) in enumerate(summary_data, 1):
            summary_ws.cell(row=row_idx, column=1, value=col1)
            summary_ws.cell(row=row_idx, column=2, value=col2)
            
        # Format summary sheet
        title_font = Font(bold=True, size=14)
        summary_ws.cell(row=1, column=1).font = title_font
        
        # Auto-adjust column widths
        summary_ws.column_dimensions['A'].width = 25
        summary_ws.column_dimensions['B'].width = 30
        
    def _sanitize_filename(self, filename: str) -> str:
        """
        Sanitize filename by removing invalid characters.
        
        Args:
            filename: Original filename
            
        Returns:
            Sanitized filename
        """
        # Remove or replace invalid characters
        invalid_chars = '<>:"/\\|?*'
        for char in invalid_chars:
            filename = filename.replace(char, '_')
            
        # Replace spaces with underscores
        filename = filename.replace(' ', '_')
        
        # Limit length
        if len(filename) > 50:
            filename = filename[:50]
            
        return filename
        
    def get_export_directory(self) -> str:
        """
        Get the export directory path.
        
        Returns:
            Export directory path
        """
        return str(self.output_dir)
        
    def list_exports(self) -> List[Dict[str, Any]]:
        """
        List all exported Excel files.
        
        Returns:
            List of export file information
        """
        exports = []
        for file_path in self.output_dir.glob("*.xlsx"):
            stat = file_path.stat()
            exports.append({
                "filename": file_path.name,
                "path": str(file_path),
                "size_bytes": stat.st_size,
                "created": datetime.fromtimestamp(stat.st_ctime).isoformat(),
                "modified": datetime.fromtimestamp(stat.st_mtime).isoformat()
            })
            
        return sorted(exports, key=lambda x: x["modified"], reverse=True) 